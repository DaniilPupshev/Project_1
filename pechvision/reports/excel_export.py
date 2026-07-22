import argparse
from collections import Counter, defaultdict
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from statistics import median
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PillowImage
from sqlalchemy import select
from sqlalchemy.orm import Session

from pechvision.config.loader import load_config
from pechvision.db.models import (
    Face,
    Person,
    ProcessingRun,
    Receipt,
    ReceiptMatch,
    Video,
    Visit,
)
from pechvision.db.session import make_engine, make_session_factory

TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
SECTION_FILL = PatternFill('solid', fgColor='D9EAF7')
HEADER_FILL = PatternFill('solid', fgColor='5B9BD5')
SUCCESS_FILL = PatternFill('solid', fgColor='E2F0D9')
WARNING_FILL = PatternFill('solid', fgColor='FFF2CC')
ERROR_FILL = PatternFill('solid', fgColor='FCE4D6')
WHITE_FONT = Font(color='FFFFFF', bold=True)
HEADER_FONT = Font(color='FFFFFF', bold=True)
THIN_GRAY = Side(style='thin', color='D9E1F2')
CELL_BORDER = Border(
    left=THIN_GRAY,
    right=THIN_GRAY,
    top=THIN_GRAY,
    bottom=THIN_GRAY,
)
PHOTO_SIZE = (96, 96)
MAX_RECEIPT_PHOTOS = 5


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description='Формирует управленческий Excel-отчёт PechVision.',
    )
    parser.add_argument('config_path', help='Путь к YAML-конфигурации')
    parser.add_argument(
        '--output',
        help='Путь к итоговому XLSX-файлу',
    )
    return parser.parse_args()


def as_local_excel_datetime(
    value: datetime | None,
    timezone: ZoneInfo,
) -> datetime | None:
    if value is None:
        return None

    if value.tzinfo is None:
        return value

    return value.astimezone(timezone).replace(tzinfo=None)


def russian_gender(value: str | None) -> str:
    return {
        'male': 'Мужчина',
        'female': 'Женщина',
    }.get(value or '', 'Не определено')


def yes_no(value: bool) -> str:
    return 'Да' if value else 'Нет'


def visit_dating_method(visit: Visit) -> str:
    if visit.entered_at is None or visit.left_at is None:
        return 'Не определено'

    if visit.time_is_estimated:
        return 'Восстановлено'

    return 'OCR'


def select_best_face(faces: list[Face]) -> Face | None:
    if not faces:
        return None

    return max(
        faces,
        key=lambda face: (
            bool(face.is_best),
            bool(face.is_identity_reference),
            float(face.identity_quality_score or 0.0),
            float(face.quality_score or 0.0),
            -(face.id or 0),
        ),
    )


def build_person_profile(
    person: Person,
    faces: list[Face],
) -> dict[str, Any]:
    reference_faces = [
        face
        for face in faces
        if face.is_identity_reference
    ]
    profile_faces = reference_faces or [
        face
        for face in faces
        if face.is_identity_eligible
    ] or faces

    gender_faces = [
        face
        for face in profile_faces
        if face.gender in {'male', 'female'}
    ]
    gender = None

    if gender_faces:
        votes = Counter(face.gender for face in gender_faces)
        max_votes = max(votes.values())
        candidates = {
            value
            for value, count in votes.items()
            if count == max_votes
        }
        gender = max(
            candidates,
            key=lambda candidate: sum(
                float(face.identity_quality_score or 0.0)
                for face in gender_faces
                if face.gender == candidate
            ),
        )

    ages = [
        float(face.age_estimate)
        for face in profile_faces
        if face.age_estimate is not None
    ]
    best_face = select_best_face(profile_faces)

    return {
        'person': person,
        'gender': gender,
        'age': median(ages) if ages else None,
        'best_face': best_face,
        'reference_faces': reference_faces,
    }


def infer_video_days(
    videos: list[Video],
    visits: list[Visit],
) -> dict[int, date | None]:
    dates_by_video: dict[int, list[date]] = defaultdict(list)

    for visit in visits:
        if visit.visit_date is not None:
            dates_by_video[visit.video_id].append(visit.visit_date)

    result: dict[int, date | None] = {}

    for video in videos:
        known_dates = dates_by_video.get(video.id, [])

        if known_dates:
            result[video.id] = min(known_dates)
        elif video.recorded_start_at is not None:
            result[video.id] = video.recorded_start_at.date()
        else:
            result[video.id] = None

    return result


def resolve_visit_day(
    visit: Visit,
    video_days: dict[int, date | None],
) -> date | None:
    return visit.visit_date or video_days.get(visit.video_id)


def create_thumbnail_buffer(path: str | None) -> BytesIO | None:
    if not path:
        return None

    image_path = Path(path)

    if not image_path.is_file():
        return None

    with PillowImage.open(image_path) as source:
        image = source.convert('RGB')
        image.thumbnail(PHOTO_SIZE)
        buffer = BytesIO()
        image.save(buffer, format='JPEG', quality=78, optimize=True)
        buffer.seek(0)

    return buffer


def insert_photo(
    worksheet: Any,
    cell_coordinate: str,
    image_path: str | None,
    image_buffers: list[BytesIO],
) -> bool:
    buffer = create_thumbnail_buffer(image_path)

    if buffer is None:
        worksheet[cell_coordinate] = 'Нет фото'
        return False

    image_buffers.append(buffer)
    image = ExcelImage(buffer)
    image.width = PHOTO_SIZE[0]
    image.height = PHOTO_SIZE[1]
    worksheet.add_image(image, cell_coordinate)
    return True


def add_sheet_title(
    worksheet: Any,
    title: str,
    subtitle: str,
    columns_count: int,
) -> None:
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=columns_count,
    )
    title_cell = worksheet.cell(1, 1, title)
    title_cell.fill = TITLE_FILL
    title_cell.font = Font(color='FFFFFF', bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    worksheet.row_dimensions[1].height = 28

    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=columns_count,
    )
    subtitle_cell = worksheet.cell(2, 1, subtitle)
    subtitle_cell.font = Font(color='666666', italic=True, size=10)
    subtitle_cell.alignment = Alignment(wrap_text=True, vertical='center')
    worksheet.row_dimensions[2].height = 28


def write_headers(
    worksheet: Any,
    row: int,
    headers: list[str],
) -> None:
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row, column, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True,
        )
        cell.border = CELL_BORDER

    worksheet.row_dimensions[row].height = 36


def style_data_range(
    worksheet: Any,
    start_row: int,
    end_row: int,
    columns_count: int,
) -> None:
    for row in worksheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=1,
        max_col=columns_count,
    ):
        for cell in row:
            cell.border = CELL_BORDER
            cell.alignment = Alignment(
                vertical='center',
                wrap_text=True,
            )


def set_widths(
    worksheet: Any,
    widths: dict[int, float],
) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = width


def build_report_data(session: Session) -> dict[str, Any]:
    videos = list(session.scalars(select(Video).order_by(Video.id)))
    runs = list(session.scalars(select(ProcessingRun).order_by(ProcessingRun.id)))
    visits = list(session.scalars(select(Visit).order_by(Visit.video_id, Visit.id)))
    faces = list(session.scalars(select(Face).order_by(Face.id)))
    persons = list(session.scalars(select(Person).order_by(Person.id)))
    receipts = list(session.scalars(select(Receipt).order_by(Receipt.closed_at, Receipt.id)))
    matches = list(session.scalars(select(ReceiptMatch).order_by(ReceiptMatch.id)))

    faces_by_visit: dict[int, list[Face]] = defaultdict(list)
    faces_by_person: dict[int, list[Face]] = defaultdict(list)

    for face in faces:
        faces_by_visit[face.visit_id].append(face)

        if face.person_id is not None:
            faces_by_person[face.person_id].append(face)

    visit_faces = {
        visit_id: select_best_face(visit_face_list)
        for visit_id, visit_face_list in faces_by_visit.items()
    }
    person_profiles = {
        person.id: build_person_profile(
            person=person,
            faces=faces_by_person.get(person.id, []),
        )
        for person in persons
    }

    visits_by_id = {visit.id: visit for visit in visits}
    receipts_by_id = {receipt.id: receipt for receipt in receipts}
    videos_by_id = {video.id: video for video in videos}
    matches_by_receipt: dict[int, list[ReceiptMatch]] = defaultdict(list)
    matches_by_visit: dict[int, list[ReceiptMatch]] = defaultdict(list)

    for match in matches:
        matches_by_receipt[match.receipt_id].append(match)
        matches_by_visit[match.visit_id].append(match)

    return {
        'videos': videos,
        'runs': runs,
        'visits': visits,
        'faces': faces,
        'persons': persons,
        'receipts': receipts,
        'matches': matches,
        'visit_faces': visit_faces,
        'person_profiles': person_profiles,
        'visits_by_id': visits_by_id,
        'receipts_by_id': receipts_by_id,
        'videos_by_id': videos_by_id,
        'matches_by_receipt': matches_by_receipt,
        'matches_by_visit': matches_by_visit,
        'video_days': infer_video_days(videos, visits),
    }


def calculate_daily_metrics(data: dict[str, Any]) -> list[dict[str, Any]]:
    visits: list[Visit] = data['visits']
    receipts: list[Receipt] = data['receipts']
    matches: list[ReceiptMatch] = data['matches']
    profiles: dict[int, dict[str, Any]] = data['person_profiles']
    receipts_by_id: dict[int, Receipt] = data['receipts_by_id']
    video_days: dict[int, date | None] = data['video_days']

    visits_by_day: dict[date, list[Visit]] = defaultdict(list)
    receipts_by_day: dict[date, list[Receipt]] = defaultdict(list)
    matches_by_day: dict[date, list[ReceiptMatch]] = defaultdict(list)

    for visit in visits:
        day = resolve_visit_day(visit, video_days)

        if day is not None:
            visits_by_day[day].append(visit)

    for receipt in receipts:
        receipts_by_day[receipt.closed_at.date()].append(receipt)

    for match in matches:
        receipt = receipts_by_id.get(match.receipt_id)

        if receipt is not None:
            matches_by_day[receipt.closed_at.date()].append(match)

    days = sorted(set(visits_by_day) | set(receipts_by_day))
    result = []

    for day in days:
        day_visits = visits_by_day.get(day, [])
        day_receipts = receipts_by_day.get(day, [])
        day_matches = matches_by_day.get(day, [])
        person_ids = {
            visit.person_id
            for visit in day_visits
            if visit.person_id is not None
        }
        genders = [
            profiles[person_id]['gender']
            for person_id in person_ids
            if person_id in profiles
        ]
        ages = [
            float(profiles[person_id]['age'])
            for person_id in person_ids
            if person_id in profiles and profiles[person_id]['age'] is not None
        ]
        matched_receipt_ids = {
            match.receipt_id
            for match in day_matches
        }
        matched_receipts = [
            receipts_by_id[receipt_id]
            for receipt_id in matched_receipt_ids
            if receipt_id in receipts_by_id
        ]

        result.append({
            'day': day,
            'visits': len(day_visits),
            'timed_visits': sum(
                visit.entered_at is not None and visit.left_at is not None
                for visit in day_visits
            ),
            'estimated_visits': sum(
                visit.entered_at is not None
                and visit.left_at is not None
                and visit.time_is_estimated
                for visit in day_visits
            ),
            'missing_time_visits': sum(
                visit.entered_at is None or visit.left_at is None
                for visit in day_visits
            ),
            'identified_visits': sum(
                visit.person_id is not None
                for visit in day_visits
            ),
            'unique_persons': len(person_ids),
            'male': genders.count('male'),
            'female': genders.count('female'),
            'unknown_gender': len(person_ids) - len([
                gender
                for gender in genders
                if gender in {'male', 'female'}
            ]),
            'average_age': sum(ages) / len(ages) if ages else None,
            'unknown_age': len(person_ids) - len(ages),
            'receipts': len(day_receipts),
            'matched_receipts': len(matched_receipt_ids),
            'coverage': (
                100.0 * len(matched_receipt_ids) / len(day_receipts)
                if day_receipts
                else 0.0
            ),
            'revenue': sum(
                (receipt.amount or Decimal('0'))
                for receipt in day_receipts
            ),
            'matched_revenue': sum(
                (receipt.amount or Decimal('0'))
                for receipt in matched_receipts
            ),
            'ambiguous_matches': sum(
                match.is_ambiguous
                for match in day_matches
            ),
        })

    return result


def calculate_total_metrics(
    data: dict[str, Any],
    daily_metrics: list[dict[str, Any]],
) -> dict[str, Any]:
    visits: list[Visit] = data['visits']
    receipts: list[Receipt] = data['receipts']
    matches: list[ReceiptMatch] = data['matches']
    profiles: dict[int, dict[str, Any]] = data['person_profiles']
    receipts_by_id: dict[int, Receipt] = data['receipts_by_id']
    person_ids = {
        visit.person_id
        for visit in visits
        if visit.person_id is not None
    }
    genders = [
        profiles[person_id]['gender']
        for person_id in person_ids
        if person_id in profiles
    ]
    ages = [
        float(profiles[person_id]['age'])
        for person_id in person_ids
        if person_id in profiles and profiles[person_id]['age'] is not None
    ]
    matched_receipt_ids = {match.receipt_id for match in matches}
    receipt_match_counts = Counter(match.receipt_id for match in matches)
    timed_visits = [
        visit
        for visit in visits
        if visit.entered_at is not None and visit.left_at is not None
    ]
    durations = [
        float(visit.duration_seconds)
        for visit in visits
        if visit.duration_seconds is not None
    ]

    return {
        'period_start': min(
            (row['day'] for row in daily_metrics),
            default=None,
        ),
        'period_end': max(
            (row['day'] for row in daily_metrics),
            default=None,
        ),
        'videos': len(data['videos']),
        'finished_runs': sum(run.status == 'finished' for run in data['runs']),
        'visits': len(visits),
        'timed_visits': len(timed_visits),
        'estimated_timed_visits': sum(
            visit.time_is_estimated
            for visit in timed_visits
        ),
        'identified_visits': sum(
            visit.person_id is not None
            for visit in visits
        ),
        'unique_persons': len(person_ids),
        'male': genders.count('male'),
        'female': genders.count('female'),
        'unknown_gender': len(person_ids) - len([
            gender
            for gender in genders
            if gender in {'male', 'female'}
        ]),
        'average_age': sum(ages) / len(ages) if ages else None,
        'unknown_age': len(person_ids) - len(ages),
        'receipts': len(receipts),
        'matched_receipts': len(matched_receipt_ids),
        'match_rows': len(matches),
        'coverage': (
            100.0 * len(matched_receipt_ids) / len(receipts)
            if receipts
            else 0.0
        ),
        'revenue': sum(
            (receipt.amount or Decimal('0'))
            for receipt in receipts
        ),
        'matched_revenue': sum(
            (receipts_by_id[receipt_id].amount or Decimal('0'))
            for receipt_id in matched_receipt_ids
            if receipt_id in receipts_by_id
        ),
        'average_duration': (
            sum(durations) / len(durations)
            if durations
            else None
        ),
        'ambiguous_matches': sum(match.is_ambiguous for match in matches),
        'receipts_with_multiple_visits': sum(
            count > 1
            for count in receipt_match_counts.values()
        ),
    }


def build_general_sheet(
    workbook: Workbook,
    data: dict[str, Any],
    daily_metrics: list[dict[str, Any]],
    total: dict[str, Any],
    timezone: ZoneInfo,
) -> None:
    worksheet = workbook.active
    worksheet.title = 'Общая сводка'
    add_sheet_title(
        worksheet,
        'PechVision — общая сводка',
        'Краткий управленческий отчёт по посещениям кассовой зоны и чекам.',
        8,
    )
    worksheet.sheet_view.showGridLines = False
    period = 'Не определён'

    if total['period_start'] and total['period_end']:
        period = (
            f'{total["period_start"]:%d.%m.%Y} — '
            f'{total["period_end"]:%d.%m.%Y}'
        )

    worksheet['A4'] = 'Период отчёта'
    worksheet['B4'] = period
    worksheet['D4'] = 'Сформирован'
    worksheet['E4'] = datetime.now(timezone).replace(tzinfo=None)
    worksheet['E4'].number_format = 'DD.MM.YYYY HH:MM'

    worksheet.merge_cells('A6:H6')
    worksheet['A6'] = 'Ключевые показатели'
    worksheet['A6'].fill = SECTION_FILL
    worksheet['A6'].font = Font(bold=True, size=12)

    indicators = [
        ('Обработано видео', total['videos']),
        ('Успешных запусков', total['finished_runs']),
        ('Всего визитов', total['visits']),
        ('Визитов со временем', total['timed_visits']),
        ('Идентифицировано визитов', total['identified_visits']),
        ('Уникальных персон', total['unique_persons']),
        ('Всего чеков', total['receipts']),
        ('Сопоставлено чеков', total['matched_receipts']),
        ('Покрытие чеков', total['coverage'] / 100),
        ('Средняя длительность визита, сек.', total['average_duration']),
        ('Общая выручка', total['revenue']),
        ('Сопоставленная выручка', total['matched_revenue']),
        ('Неоднозначных совпадений', total['ambiguous_matches']),
        ('Чеков с несколькими визитами', total['receipts_with_multiple_visits']),
    ]

    for index, (label, value) in enumerate(indicators):
        row = 7 + index // 2
        column = 1 if index % 2 == 0 else 5
        worksheet.cell(row, column, label).font = Font(bold=True)
        value_cell = worksheet.cell(row, column + 1, value)
        value_cell.fill = SUCCESS_FILL
        value_cell.font = Font(bold=True, size=12)
        value_cell.alignment = Alignment(horizontal='center')

        if 'Покрытие' in label:
            value_cell.number_format = '0.0%'
        elif 'выручка' in label.lower():
            value_cell.number_format = '#,##0.00 [$₽-ru-RU]'
        elif isinstance(value, float):
            value_cell.number_format = '0.0'

    table_row = 16
    headers = [
        'Дата',
        'Визитов',
        'Уникальных персон',
        'Чеков',
        'Совпадений',
        'Покрытие, %',
        'Выручка',
        'Сопоставленная выручка',
    ]
    write_headers(worksheet, table_row, headers)

    for row_index, metric in enumerate(daily_metrics, start=table_row + 1):
        values = [
            metric['day'],
            metric['visits'],
            metric['unique_persons'],
            metric['receipts'],
            metric['matched_receipts'],
            metric['coverage'] / 100,
            metric['revenue'],
            metric['matched_revenue'],
        ]

        for column, value in enumerate(values, start=1):
            worksheet.cell(row_index, column, value)

        worksheet.cell(row_index, 1).number_format = 'DD.MM.YYYY'
        worksheet.cell(row_index, 6).number_format = '0.0%'
        worksheet.cell(row_index, 7).number_format = '#,##0.00 [$₽-ru-RU]'
        worksheet.cell(row_index, 8).number_format = '#,##0.00 [$₽-ru-RU]'

    end_row = table_row + len(daily_metrics)
    style_data_range(worksheet, table_row + 1, end_row, len(headers))
    worksheet.auto_filter.ref = f'A{table_row}:H{end_row}'
    worksheet.freeze_panes = f'A{table_row + 1}'

    if daily_metrics:
        bar_chart = BarChart()
        bar_chart.title = 'Чеки и совпадения по дням'
        bar_chart.y_axis.title = 'Количество'
        bar_chart.x_axis.title = 'Дата'
        bar_data = Reference(
            worksheet,
            min_col=4,
            max_col=5,
            min_row=table_row,
            max_row=end_row,
        )
        categories = Reference(
            worksheet,
            min_col=1,
            min_row=table_row + 1,
            max_row=end_row,
        )
        bar_chart.add_data(bar_data, titles_from_data=True)
        bar_chart.set_categories(categories)
        bar_chart.height = 7
        bar_chart.width = 13
        worksheet.add_chart(bar_chart, 'J4')

        line_chart = LineChart()
        line_chart.title = 'Покрытие чеков по дням'
        line_chart.y_axis.title = 'Покрытие'
        line_chart.y_axis.numFmt = '0%'
        coverage_data = Reference(
            worksheet,
            min_col=6,
            min_row=table_row,
            max_row=end_row,
        )
        line_chart.add_data(coverage_data, titles_from_data=True)
        line_chart.set_categories(categories)
        line_chart.height = 7
        line_chart.width = 13
        worksheet.add_chart(line_chart, 'J19')

    notes_row = end_row + 3
    worksheet.merge_cells(
        start_row=notes_row,
        start_column=1,
        end_row=notes_row,
        end_column=8,
    )
    worksheet.cell(notes_row, 1, 'Как читать отчёт').fill = SECTION_FILL
    worksheet.cell(notes_row, 1).font = Font(bold=True, size=12)
    notes = [
        'Совпадение — уникальный чек, для которого найден хотя бы один визит.',
        'Покрытие — доля сопоставленных уникальных чеков от всех чеков.',
        'Один чек может быть связан с несколькими визитами; сумма чека учитывается один раз.',
        'Пол и возраст являются приблизительной оценкой модели компьютерного зрения.',
        'Восстановленное время рассчитано по соседним OCR-меткам и длительности трека.',
        'Фотографии лиц предназначены только для внутреннего использования '
        'с ограниченным доступом.',
    ]

    for offset, note in enumerate(notes, start=1):
        worksheet.merge_cells(
            start_row=notes_row + offset,
            start_column=1,
            end_row=notes_row + offset,
            end_column=8,
        )
        worksheet.cell(notes_row + offset, 1, f'• {note}')
        worksheet.cell(notes_row + offset, 1).alignment = Alignment(wrap_text=True)

    set_widths(
        worksheet,
        {
            1: 23,
            2: 18,
            3: 20,
            4: 18,
            5: 24,
            6: 18,
            7: 20,
            8: 24,
        },
    )


def build_people_summary_sheet(
    workbook: Workbook,
    daily_metrics: list[dict[str, Any]],
    total: dict[str, Any],
) -> None:
    worksheet = workbook.create_sheet('Сводка по людям')
    headers = [
        'Дата',
        'Уникальных персон',
        'Муж',
        'Жен',
        'Пол не определён',
        'Средний возраст',
        'Без оценки возраста',
        'Всего визитов',
        'Идентифицировано визитов',
        'Чеков',
        'Совпадений',
        'Покрытие, %',
        'Выручка',
        'Сопоставленная выручка',
    ]
    add_sheet_title(
        worksheet,
        'Сводка по людям и чекам',
        'Демография рассчитана по уникальным персонам, а не по отдельным фотографиям.',
        len(headers),
    )
    header_row = 4
    write_headers(worksheet, header_row, headers)

    rows = [
        [
            metric['day'],
            metric['unique_persons'],
            metric['male'],
            metric['female'],
            metric['unknown_gender'],
            metric['average_age'],
            metric['unknown_age'],
            metric['visits'],
            metric['identified_visits'],
            metric['receipts'],
            metric['matched_receipts'],
            metric['coverage'] / 100,
            metric['revenue'],
            metric['matched_revenue'],
        ]
        for metric in daily_metrics
    ]
    rows.append([
        'ИТОГО',
        total['unique_persons'],
        total['male'],
        total['female'],
        total['unknown_gender'],
        total['average_age'],
        total['unknown_age'],
        total['visits'],
        total['identified_visits'],
        total['receipts'],
        total['matched_receipts'],
        total['coverage'] / 100,
        total['revenue'],
        total['matched_revenue'],
    ])

    for row_index, values in enumerate(rows, start=header_row + 1):
        for column, value in enumerate(values, start=1):
            worksheet.cell(row_index, column, value)

        if isinstance(values[0], date):
            worksheet.cell(row_index, 1).number_format = 'DD.MM.YYYY'
        else:
            for cell in worksheet[row_index]:
                cell.font = Font(bold=True)
                cell.fill = SECTION_FILL

        worksheet.cell(row_index, 6).number_format = '0.0'
        worksheet.cell(row_index, 12).number_format = '0.0%'
        worksheet.cell(row_index, 13).number_format = '#,##0.00 [$₽-ru-RU]'
        worksheet.cell(row_index, 14).number_format = '#,##0.00 [$₽-ru-RU]'

    end_row = header_row + len(rows)
    style_data_range(worksheet, header_row + 1, end_row, len(headers))
    worksheet.auto_filter.ref = f'A{header_row}:N{end_row}'
    worksheet.freeze_panes = f'A{header_row + 1}'
    set_widths(
        worksheet,
        {column: 18 for column in range(1, len(headers) + 1)},
    )
    worksheet.column_dimensions['A'].width = 14
    worksheet.column_dimensions['I'].width = 24
    worksheet.column_dimensions['N'].width = 25


def build_receipts_sheet(
    workbook: Workbook,
    data: dict[str, Any],
    timezone: ZoneInfo,
    image_buffers: list[BytesIO],
) -> None:
    worksheet = workbook.create_sheet('Сводка по чекам')
    photo_headers = [
        f'Фото {index}'
        for index in range(1, MAX_RECEIPT_PHOTOS + 1)
    ]
    headers = [
        'ID чека',
        'Дата',
        'Открытие',
        'Закрытие',
        'Обслуживание, мин.',
        'Сумма',
        'Торговая точка',
        'Номер стола',
        'Статус',
        'Связанных визитов',
        'ID персон',
        'ID визитов',
        'Способ датирования',
        'Мин. разница, сек.',
        'Неоднозначное совпадение',
        'Групповое совпадение',
        *photo_headers,
        'Комментарий',
        'Источник чеков',
    ]
    add_sheet_title(
        worksheet,
        'Сводка по чекам',
        'Одна строка соответствует одному чеку. Сумма не дублируется при нескольких посетителях.',
        len(headers),
    )
    header_row = 4
    write_headers(worksheet, header_row, headers)
    visits_by_id: dict[int, Visit] = data['visits_by_id']
    matches_by_receipt: dict[int, list[ReceiptMatch]] = data['matches_by_receipt']
    visit_faces: dict[int, Face | None] = data['visit_faces']

    for row_index, receipt in enumerate(data['receipts'], start=header_row + 1):
        receipt_matches = matches_by_receipt.get(receipt.id, [])
        matched_visits = [
            visits_by_id[match.visit_id]
            for match in receipt_matches
            if match.visit_id in visits_by_id
        ]
        opened_at = as_local_excel_datetime(receipt.opened_at, timezone)
        closed_at = as_local_excel_datetime(receipt.closed_at, timezone)
        service_minutes = None

        if opened_at is not None and closed_at is not None:
            service_minutes = (closed_at - opened_at).total_seconds() / 60

        person_ids = sorted({
            visit.person_id
            for visit in matched_visits
            if visit.person_id is not None
        })
        dating_methods = sorted({
            visit_dating_method(visit)
            for visit in matched_visits
        })
        values = [
            receipt.external_receipt_id,
            closed_at.date() if closed_at else None,
            opened_at,
            closed_at,
            service_minutes,
            receipt.amount,
            receipt.tt,
            receipt.table_number,
            'Сопоставлен' if receipt_matches else 'Не сопоставлен',
            len(matched_visits),
            ', '.join(map(str, person_ids)) or 'Не определено',
            ', '.join(str(visit.id) for visit in matched_visits) or 'Не определено',
            ', '.join(dating_methods) or 'Не определено',
            min(
                (match.time_delta_seconds for match in receipt_matches),
                default=None,
            ),
            yes_no(any(match.is_ambiguous for match in receipt_matches)),
            yes_no(
                any(match.is_group_match for match in receipt_matches)
                or len(matched_visits) > 1
            ),
        ]

        for column, value in enumerate(values, start=1):
            worksheet.cell(row_index, column, value)

        photo_start_column = 17

        for photo_index in range(MAX_RECEIPT_PHOTOS):
            column = photo_start_column + photo_index

            if photo_index >= len(matched_visits):
                worksheet.cell(row_index, column, '—')
                continue

            visit = matched_visits[photo_index]
            face = visit_faces.get(visit.id)
            insert_photo(
                worksheet=worksheet,
                cell_coordinate=f'{get_column_letter(column)}{row_index}',
                image_path=face.image_path if face else None,
                image_buffers=image_buffers,
            )

        comment_column = photo_start_column + MAX_RECEIPT_PHOTOS
        source_column = comment_column + 1
        comments = []

        if len(matched_visits) > MAX_RECEIPT_PHOTOS:
            comments.append('Показаны не все фотографии')

        if len(matched_visits) > 1:
            comments.append('Один чек связан с несколькими визитами')

        if any(match.is_ambiguous for match in receipt_matches):
            comments.append('Было несколько чеков-кандидатов')

        worksheet.cell(
            row_index,
            comment_column,
            '; '.join(comments) or '—',
        )
        worksheet.cell(row_index, source_column, receipt.source_file)
        worksheet.row_dimensions[row_index].height = 78 if matched_visits else 24
        worksheet.cell(row_index, 2).number_format = 'DD.MM.YYYY'
        worksheet.cell(row_index, 3).number_format = 'DD.MM.YYYY HH:MM:SS'
        worksheet.cell(row_index, 4).number_format = 'DD.MM.YYYY HH:MM:SS'
        worksheet.cell(row_index, 5).number_format = '0.0'
        worksheet.cell(row_index, 6).number_format = '#,##0.00 [$₽-ru-RU]'
        worksheet.cell(row_index, 14).number_format = '0.0'
        status_cell = worksheet.cell(row_index, 9)
        status_cell.fill = SUCCESS_FILL if receipt_matches else WARNING_FILL

    end_row = header_row + len(data['receipts'])
    style_data_range(worksheet, header_row + 1, end_row, len(headers))
    worksheet.auto_filter.ref = (
        f'A{header_row}:{get_column_letter(len(headers))}{end_row}'
    )
    worksheet.freeze_panes = f'A{header_row + 1}'
    widths = {
        1: 38,
        2: 13,
        3: 21,
        4: 21,
        5: 18,
        6: 14,
        7: 18,
        8: 14,
        9: 18,
        10: 18,
        11: 22,
        12: 22,
        13: 24,
        14: 18,
        15: 20,
        16: 20,
        comment_column: 34,
        source_column: 38,
    }

    for column in range(photo_start_column, comment_column):
        widths[column] = 15

    set_widths(worksheet, widths)


def build_visits_sheet(
    workbook: Workbook,
    data: dict[str, Any],
    timezone: ZoneInfo,
    image_buffers: list[BytesIO],
) -> None:
    worksheet = workbook.create_sheet('Сводка по визитам')
    headers = [
        'ID визита',
        'ID персоны',
        'Фото в момент визита',
        'Дата визита',
        'Дата определена по видео',
        'Время входа',
        'Время выхода',
        'Длительность, сек.',
        'Кадр входа',
        'Кадр выхода',
        'Способ датирования',
        'OCR на входе',
        'OCR на выходе',
        'Причина отклонения OCR',
        'ID видео',
        'Имя видео',
        'ID трека',
        'Лицо пригодно',
        'Качество лица',
        'Пол',
        'Примерный возраст',
        'ID чека',
        'Сумма чека',
        'Разница по времени, сек.',
        'Неоднозначное совпадение',
        'Персонал',
        'Группа',
    ]
    add_sheet_title(
        worksheet,
        'Сводка по визитам',
        'Одна строка соответствует одному зафиксированному визиту в кассовой зоне.',
        len(headers),
    )
    header_row = 4
    write_headers(worksheet, header_row, headers)
    videos_by_id: dict[int, Video] = data['videos_by_id']
    receipts_by_id: dict[int, Receipt] = data['receipts_by_id']
    matches_by_visit: dict[int, list[ReceiptMatch]] = data['matches_by_visit']
    visit_faces: dict[int, Face | None] = data['visit_faces']
    video_days: dict[int, date | None] = data['video_days']

    for row_index, visit in enumerate(data['visits'], start=header_row + 1):
        face = visit_faces.get(visit.id)
        video = videos_by_id.get(visit.video_id)
        visit_matches = matches_by_visit.get(visit.id, [])
        linked_receipts = [
            receipts_by_id[match.receipt_id]
            for match in visit_matches
            if match.receipt_id in receipts_by_id
        ]
        extra_data = visit.extra_data or {}
        inferred_day = video_days.get(visit.video_id)
        display_day = visit.visit_date or inferred_day
        values = [
            visit.id,
            visit.person_id or 'Не определено',
            None,
            display_day,
            yes_no(visit.visit_date is None and inferred_day is not None),
            as_local_excel_datetime(visit.entered_at, timezone),
            as_local_excel_datetime(visit.left_at, timezone),
            visit.duration_seconds,
            visit.entry_frame_index,
            visit.exit_frame_index,
            visit_dating_method(visit),
            extra_data.get('ocr_entry_text') or 'Не определено',
            extra_data.get('ocr_exit_text') or 'Не определено',
            extra_data.get('ocr_rejection_reason') or '—',
            visit.video_id,
            video.filename if video else 'Не определено',
            visit.track_id,
            yes_no(bool(face and face.is_identity_eligible)),
            face.identity_quality_score if face else None,
            russian_gender(face.gender if face else None),
            face.age_estimate if face else None,
            ', '.join(
                receipt.external_receipt_id
                for receipt in linked_receipts
            ) or 'Не определено',
            sum(
                (receipt.amount or Decimal('0'))
                for receipt in {
                    receipt.id: receipt
                    for receipt in linked_receipts
                }.values()
            ) if linked_receipts else None,
            min(
                (match.time_delta_seconds for match in visit_matches),
                default=None,
            ),
            yes_no(any(match.is_ambiguous for match in visit_matches)),
            yes_no(visit.is_staff),
            yes_no(visit.is_group),
        ]

        for column, value in enumerate(values, start=1):
            worksheet.cell(row_index, column, value)

        insert_photo(
            worksheet=worksheet,
            cell_coordinate=f'C{row_index}',
            image_path=face.image_path if face else None,
            image_buffers=image_buffers,
        )
        worksheet.row_dimensions[row_index].height = 78 if face else 24
        worksheet.cell(row_index, 4).number_format = 'DD.MM.YYYY'
        worksheet.cell(row_index, 6).number_format = 'DD.MM.YYYY HH:MM:SS'
        worksheet.cell(row_index, 7).number_format = 'DD.MM.YYYY HH:MM:SS'
        worksheet.cell(row_index, 8).number_format = '0.0'
        worksheet.cell(row_index, 19).number_format = '0.000'
        worksheet.cell(row_index, 21).number_format = '0'
        worksheet.cell(row_index, 23).number_format = '#,##0.00 [$₽-ru-RU]'
        worksheet.cell(row_index, 24).number_format = '0.0'

        dating_cell = worksheet.cell(row_index, 11)

        if dating_cell.value == 'OCR':
            dating_cell.fill = SUCCESS_FILL
        elif dating_cell.value == 'Восстановлено':
            dating_cell.fill = WARNING_FILL
        else:
            dating_cell.fill = ERROR_FILL

    end_row = header_row + len(data['visits'])
    style_data_range(worksheet, header_row + 1, end_row, len(headers))
    worksheet.auto_filter.ref = (
        f'A{header_row}:{get_column_letter(len(headers))}{end_row}'
    )
    worksheet.freeze_panes = f'A{header_row + 1}'
    widths = {column: 17 for column in range(1, len(headers) + 1)}
    widths.update({
        3: 16,
        5: 23,
        6: 21,
        7: 21,
        11: 20,
        12: 30,
        13: 30,
        14: 30,
        16: 24,
        22: 38,
        25: 22,
    })
    set_widths(worksheet, widths)


def build_person_registry_sheet(
    workbook: Workbook,
    data: dict[str, Any],
    timezone: ZoneInfo,
    image_buffers: list[BytesIO],
) -> None:
    worksheet = workbook.create_sheet('Реестр персон')
    headers = [
        'ID персоны',
        'Лучшее фото',
        'Предполагаемый пол',
        'Примерный возраст',
        'Первое появление',
        'Последнее появление',
        'Дней посещения',
        'Количество визитов',
        'Связанных чеков',
        'Сумма уникальных чеков',
        'Эталонных лиц',
        'Ракурсы эталонов',
        'Внешний ключ',
    ]
    add_sheet_title(
        worksheet,
        'Реестр уникальных персон',
        'Одна строка соответствует одной личности по результатам распознавания лиц.',
        len(headers),
    )
    header_row = 4
    write_headers(worksheet, header_row, headers)
    visits_by_person: dict[int, list[Visit]] = defaultdict(list)
    video_days: dict[int, date | None] = data['video_days']
    matches_by_visit: dict[int, list[ReceiptMatch]] = data['matches_by_visit']
    receipts_by_id: dict[int, Receipt] = data['receipts_by_id']

    for visit in data['visits']:
        if visit.person_id is not None:
            visits_by_person[visit.person_id].append(visit)

    for row_index, person in enumerate(data['persons'], start=header_row + 1):
        profile = data['person_profiles'][person.id]
        person_visits = visits_by_person.get(person.id, [])
        days = {
            resolve_visit_day(visit, video_days)
            for visit in person_visits
            if resolve_visit_day(visit, video_days) is not None
        }
        receipt_ids = {
            match.receipt_id
            for visit in person_visits
            for match in matches_by_visit.get(visit.id, [])
        }
        reference_faces: list[Face] = profile['reference_faces']
        poses = sorted({
            face.pose_category or 'unknown'
            for face in reference_faces
        })
        best_face: Face | None = profile['best_face']
        best_path = (
            best_face.image_path
            if best_face is not None
            else person.best_face_path
        )
        values = [
            person.id,
            None,
            russian_gender(profile['gender']),
            profile['age'],
            as_local_excel_datetime(person.first_seen_at, timezone),
            as_local_excel_datetime(person.last_seen_at, timezone),
            len(days),
            len(person_visits),
            len(receipt_ids),
            sum(
                (receipts_by_id[receipt_id].amount or Decimal('0'))
                for receipt_id in receipt_ids
                if receipt_id in receipts_by_id
            ),
            len(reference_faces),
            ', '.join(poses) or 'Не определено',
            person.external_person_key,
        ]

        for column, value in enumerate(values, start=1):
            worksheet.cell(row_index, column, value)

        insert_photo(
            worksheet=worksheet,
            cell_coordinate=f'B{row_index}',
            image_path=best_path,
            image_buffers=image_buffers,
        )
        worksheet.row_dimensions[row_index].height = 78
        worksheet.cell(row_index, 4).number_format = '0.0'
        worksheet.cell(row_index, 5).number_format = 'DD.MM.YYYY HH:MM:SS'
        worksheet.cell(row_index, 6).number_format = 'DD.MM.YYYY HH:MM:SS'
        worksheet.cell(row_index, 10).number_format = '#,##0.00 [$₽-ru-RU]'

    end_row = header_row + len(data['persons'])
    style_data_range(worksheet, header_row + 1, end_row, len(headers))
    worksheet.auto_filter.ref = f'A{header_row}:M{end_row}'
    worksheet.freeze_panes = f'A{header_row + 1}'
    set_widths(
        worksheet,
        {
            1: 14,
            2: 16,
            3: 22,
            4: 20,
            5: 22,
            6: 22,
            7: 18,
            8: 20,
            9: 18,
            10: 24,
            11: 18,
            12: 24,
            13: 38,
        },
    )


def build_quality_sheet(
    workbook: Workbook,
    data: dict[str, Any],
    daily_metrics: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet('Качество данных')
    headers = [
        'Дата',
        'Всего визитов',
        'Время определено',
        'Время восстановлено',
        'Время отсутствует',
        'Лицо найдено',
        'Лицо пригодно',
        'Визит идентифицирован',
        'Неоднозначных совпадений',
        'Чеков',
        'Совпадений',
        'Покрытие, %',
        'Оценка качества',
    ]
    add_sheet_title(
        worksheet,
        'Качество исходных данных и обработки',
        'Лист показывает полноту OCR, распознавания лиц и сопоставления с чеками.',
        len(headers),
    )
    header_row = 4
    write_headers(worksheet, header_row, headers)
    video_days: dict[int, date | None] = data['video_days']
    visit_faces: dict[int, Face | None] = data['visit_faces']
    faces_found_by_day: Counter[date] = Counter()
    eligible_by_day: Counter[date] = Counter()

    for visit in data['visits']:
        day = resolve_visit_day(visit, video_days)
        face = visit_faces.get(visit.id)

        if day is None or face is None:
            continue

        faces_found_by_day[day] += 1

        if face.is_identity_eligible:
            eligible_by_day[day] += 1

    for row_index, metric in enumerate(daily_metrics, start=header_row + 1):
        timed_ratio = (
            metric['timed_visits'] / metric['visits']
            if metric['visits']
            else 0.0
        )

        if timed_ratio < 0.6:
            quality = 'Требует внимания: низкая полнота времени'
        elif metric['coverage'] < 70:
            quality = 'Требует внимания: низкое покрытие чеков'
        elif metric['coverage'] < 85:
            quality = 'Приемлемо, есть пропуски'
        else:
            quality = 'Хорошо'

        values = [
            metric['day'],
            metric['visits'],
            metric['timed_visits'],
            metric['estimated_visits'],
            metric['missing_time_visits'],
            faces_found_by_day[metric['day']],
            eligible_by_day[metric['day']],
            metric['identified_visits'],
            metric['ambiguous_matches'],
            metric['receipts'],
            metric['matched_receipts'],
            metric['coverage'] / 100,
            quality,
        ]

        for column, value in enumerate(values, start=1):
            worksheet.cell(row_index, column, value)

        worksheet.cell(row_index, 1).number_format = 'DD.MM.YYYY'
        worksheet.cell(row_index, 12).number_format = '0.0%'
        quality_cell = worksheet.cell(row_index, 13)

        if quality == 'Хорошо':
            quality_cell.fill = SUCCESS_FILL
        elif quality.startswith('Приемлемо'):
            quality_cell.fill = WARNING_FILL
        else:
            quality_cell.fill = ERROR_FILL

    end_row = header_row + len(daily_metrics)
    style_data_range(worksheet, header_row + 1, end_row, len(headers))
    worksheet.auto_filter.ref = f'A{header_row}:M{end_row}'
    worksheet.freeze_panes = f'A{header_row + 1}'
    set_widths(
        worksheet,
        {
            1: 14,
            2: 18,
            3: 20,
            4: 22,
            5: 20,
            6: 17,
            7: 17,
            8: 24,
            9: 25,
            10: 14,
            11: 16,
            12: 16,
            13: 42,
        },
    )


def build_workbook(
    session: Session,
    timezone_name: str,
) -> tuple[Workbook, list[BytesIO]]:
    timezone = ZoneInfo(timezone_name)
    data = build_report_data(session)
    daily_metrics = calculate_daily_metrics(data)
    total_metrics = calculate_total_metrics(data, daily_metrics)
    workbook = Workbook()
    image_buffers: list[BytesIO] = []

    build_general_sheet(
        workbook=workbook,
        data=data,
        daily_metrics=daily_metrics,
        total=total_metrics,
        timezone=timezone,
    )
    build_people_summary_sheet(
        workbook=workbook,
        daily_metrics=daily_metrics,
        total=total_metrics,
    )
    build_receipts_sheet(
        workbook=workbook,
        data=data,
        timezone=timezone,
        image_buffers=image_buffers,
    )
    build_visits_sheet(
        workbook=workbook,
        data=data,
        timezone=timezone,
        image_buffers=image_buffers,
    )
    build_person_registry_sheet(
        workbook=workbook,
        data=data,
        timezone=timezone,
        image_buffers=image_buffers,
    )
    build_quality_sheet(
        workbook=workbook,
        data=data,
        daily_metrics=daily_metrics,
    )

    for worksheet in workbook.worksheets:
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_view.showGridLines = False

    return workbook, image_buffers


def default_output_path(config: Any) -> Path:
    timestamp = datetime.now(UTC).strftime('%Y%m%d_%H%M%S')
    return config.paths.reports_dir / f'Отчет_PechVision_{timestamp}.xlsx'


def export_report(
    config_path: str | Path,
    output_path: str | Path | None = None,
) -> Path:
    config = load_config(config_path)
    engine = make_engine(config)
    session_factory = make_session_factory(engine)
    output = Path(output_path) if output_path else default_output_path(config)
    output.parent.mkdir(parents=True, exist_ok=True)

    with session_factory() as session:
        workbook, image_buffers = build_workbook(
            session=session,
            timezone_name=config.project.timezone,
        )
        workbook.save(output)

    for buffer in image_buffers:
        buffer.close()

    return output


def main() -> None:
    args = parse_args()
    output = export_report(
        config_path=args.config_path,
        output_path=args.output,
    )
    print('EXCEL REPORT CREATED')
    print('-' * 20)
    print(f'Файл: {output}')


if __name__ == '__main__':
    main()
